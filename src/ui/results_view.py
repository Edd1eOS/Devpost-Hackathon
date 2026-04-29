import io

import pandas as pd
import streamlit as st

from models.schemas import AuditResult, AuditStatus, Severity

_STATUS_ORDER   = {AuditStatus.FLAGGED: 0, AuditStatus.UNCERTAIN: 1, AuditStatus.OK: 2}
_SEVERITY_ORDER = {Severity.HIGH: 0, Severity.MEDIUM: 1, Severity.LOW: 2, Severity.NONE: 3}

_STATUS_COLORS = {
    "FLAGGED":   ("🔴", "#ffcccc"),
    "UNCERTAIN": ("🟠", "#ffe0b2"),
    "OK":        ("🟢", "#c8e6c9"),
}

_SEVERITY_COLORS = {
    "HIGH":   "#ef9a9a",
    "MEDIUM": "#ffcc80",
    "LOW":    "#fff9c4",
    "NONE":   "#f5f5f5",
}


def _badge(text: str, color: str) -> str:
    return f'<span style="background:{color};padding:2px 8px;border-radius:4px;font-weight:bold">{text}</span>'


def _sort_key(r: AuditResult):
    return (_STATUS_ORDER[r.audit_status], _SEVERITY_ORDER[r.severity])


def _filter(results: list, status_filter: str, search: str) -> list:
    out = results
    if status_filter != "All":
        out = [r for r in out if r.audit_status.value == status_filter]
    if search:
        q = search.lower()
        out = [
            r for r in out
            if q in r.vendor.lower() or q in r.employee_id.lower() or q in r.category.lower()
        ]
    return out


def _results_to_df(results: list) -> pd.DataFrame:
    rows = []
    for r in results:
        emoji, _ = _STATUS_COLORS.get(r.audit_status.value, ("", ""))
        rows.append({
            "ID":         r.expense_id,
            "Vendor":     r.vendor,
            "Amount":     r.amount,
            "Category":   r.category,
            "Status":     f"{emoji} {r.audit_status.value}",
            "Confidence": f"{round(r.confidence * 100)}%",
            "Severity":   r.severity.value,
            "Reason":     r.reason_summary,
        })
    return pd.DataFrame(rows)


def _export_csv(results: list) -> bytes:
    rows = []
    for r in results:
        rows.append({
            "expense_id":          r.expense_id,
            "employee_id":         r.employee_id,
            "vendor":              r.vendor,
            "amount":              r.amount,
            "date":                r.date,
            "category":            r.category,
            "receipt_file":        r.receipt_file,
            "audit_status":        r.audit_status.value,
            "severity":            r.severity.value,
            "confidence":          r.confidence,
            "triggered_rules":     ", ".join(r.triggered_rules),
            "reason_summary":      r.reason_summary,
            "detailed_reason":     r.detailed_reason,
            "suggested_action":    r.suggested_action,
            "matched_receipt_file": r.matched_receipt_file,
            "duplicate_match_id":  r.duplicate_match_id or "",
            "ai_assisted":         r.ai_assisted,
            "ai_note":             r.ai_note or "",
            "generated_at":        r.generated_at,
        })
    df = pd.DataFrame(rows)
    return df.to_csv(index=False).encode("utf-8")


def _export_excel(results: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _FILL = {
        "FLAGGED":   PatternFill("solid", fgColor="FFCCCC"),
        "UNCERTAIN": PatternFill("solid", fgColor="FFE0B2"),
        "OK":        PatternFill("solid", fgColor="C8E6C9"),
    }
    _SEVERITY_FILL = {
        "HIGH":   PatternFill("solid", fgColor="EF9A9A"),
        "MEDIUM": PatternFill("solid", fgColor="FFCC80"),
        "LOW":    PatternFill("solid", fgColor="FFF9C4"),
        "NONE":   PatternFill("solid", fgColor="F5F5F5"),
    }

    headers = [
        "expense_id", "employee_id", "vendor", "amount", "date", "category",
        "receipt_file", "audit_status", "severity", "confidence",
        "triggered_rules", "reason_summary", "detailed_reason",
        "suggested_action", "matched_receipt_file", "duplicate_match_id",
        "ai_assisted", "ai_note", "generated_at",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    # Header row
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(results, 2):
        values = [
            r.expense_id, r.employee_id, r.vendor, r.amount, r.date,
            r.category, r.receipt_file, r.audit_status.value, r.severity.value,
            round(r.confidence, 3), ", ".join(r.triggered_rules),
            r.reason_summary, r.detailed_reason, r.suggested_action,
            r.matched_receipt_file, r.duplicate_match_id or "",
            r.ai_assisted, r.ai_note or "", r.generated_at,
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)

        # Color audit_status column (col 8) and severity column (col 9)
        ws.cell(row=row_idx, column=8).fill = _FILL.get(r.audit_status.value, PatternFill())
        ws.cell(row=row_idx, column=9).fill = _SEVERITY_FILL.get(r.severity.value, PatternFill())

    # Column widths
    col_widths = [10, 12, 20, 10, 12, 18, 20, 12, 10, 12,
                  30, 40, 50, 35, 20, 18, 12, 35, 22]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_detail(result: AuditResult):
    st.subheader("Detail Panel")

    emoji, color = _STATUS_COLORS.get(result.audit_status.value, ("", "#eee"))
    st.markdown(
        f"{emoji} **{result.audit_status.value}** &nbsp; "
        f"Severity: **{result.severity.value}** &nbsp; "
        f"Confidence: **{round(result.confidence * 100)}%**",
        unsafe_allow_html=True,
    )

    if result.triggered_rules:
        st.caption("Rules triggered: " + ", ".join(result.triggered_rules))

    st.divider()

    # Rule-specific comparison section
    rules = result.triggered_rules

    if "check_amount_mismatch" in rules and result.ocr_result and result.ocr_result.amount is not None:
        st.markdown("**Amount Comparison**")
        col1, col2 = st.columns(2)
        col1.metric("Reported", f"${result.amount:.2f}")
        col2.metric("Receipt (OCR)", f"${result.ocr_result.amount:.2f}")
        diff = abs(result.amount - result.ocr_result.amount)
        pct = diff / result.amount * 100 if result.amount else 0
        st.caption(f"Difference: ${diff:.2f} ({pct:.1f}%) | Threshold: 10%")

    if "check_duplicate" in rules and result.duplicate_match_id:
        st.markdown("**Duplicate Match**")
        st.info(f"Matches expense ID: **{result.duplicate_match_id}**")
        col1, col2 = st.columns(2)
        col1.markdown(f"**This expense**\n\nVendor: {result.vendor}\n\nAmount: ${result.amount:.2f}\n\nDate: {result.date}")
        col2.markdown(f"**Matching expense ({result.duplicate_match_id})**\n\n*(see full dataset)*")

    if "check_missing_receipt" in rules and result.ocr_result is None:
        st.markdown("**Missing Receipt**")
        st.error(f"Expected file: `{result.receipt_file}`\nStatus: not found in uploaded folder")

    if "check_category" in rules:
        st.markdown("**Category Check**")
        st.write(f"Declared: **{result.category}**")
        if result.ai_note:
            st.write(result.ai_note)

    if "check_suspicious_pattern" in rules:
        st.markdown("**Suspicious Pattern**")
        st.warning(f"Employee `{result.employee_id}` has multiple issues in this dataset.")

    st.divider()

    if result.ai_assisted and result.ai_note:
        st.markdown(f"🤖 **AI note:** {result.ai_note}")
        st.divider()

    st.markdown(f"**Suggested Action:** {result.suggested_action}")

    with st.expander("Full detail"):
        st.text(result.detailed_reason)


def render_results():
    results: list[AuditResult] = st.session_state.get("results", [])

    if not results:
        st.warning("No results found. Please run an audit first.")
        if st.button("New Audit"):
            _reset()
        return

    # --- Summary bar ---
    flagged   = [r for r in results if r.audit_status == AuditStatus.FLAGGED]
    uncertain = [r for r in results if r.audit_status == AuditStatus.UNCERTAIN]
    ok        = [r for r in results if r.audit_status == AuditStatus.OK]
    total_amount    = sum(r.amount for r in results)
    high_risk_amount = sum(r.amount for r in flagged)

    if not flagged and not uncertain:
        st.success("✔ All expenses passed audit")
        st.caption("0 flagged · 0 uncertain · 100% compliant")
    else:
        col1, col2, col3, col4, col5, col6 = st.columns(6)
        col1.metric("Total checked", len(results))
        col2.metric("✓ OK",          len(ok))
        col3.metric("⚑ Flagged",     len(flagged))
        col4.metric("△ Uncertain",   len(uncertain))
        col5.metric("Total reviewed", f"${total_amount:,.2f}")
        col6.metric("High-risk amount", f"${high_risk_amount:,.2f}")

    st.divider()

    # --- Filter & search ---
    col1, col2, col3 = st.columns([1, 1, 2])
    filter_status = col1.selectbox("Status", ["All", "FLAGGED", "UNCERTAIN", "OK"],
                                   key="filter_status")
    search_query  = col3.text_input("Search vendor / employee / category", key="search_query")

    filtered = sorted(_filter(results, filter_status, search_query), key=_sort_key)

    # --- Layout: table + detail panel ---
    main_col, detail_col = st.columns([2, 1])

    with main_col:
        st.subheader(f"Results ({len(filtered)} shown)")
        if not filtered:
            st.info("No results match the current filter.")
        else:
            table_df = _results_to_df(filtered)
            # Use radio-button selection column for row selection
            filtered_ids = [r.expense_id for r in filtered]
            selected_label = st.radio(
                "Select row to inspect",
                options=filtered_ids,
                index=0 if filtered_ids else None,
                key="selected_row_radio",
                horizontal=False,
                label_visibility="collapsed",
            )
            st.dataframe(table_df, use_container_width=True, hide_index=True)
            st.session_state["selected_row"] = selected_label

    with detail_col:
        selected_id = st.session_state.get("selected_row")
        if selected_id:
            selected_result = next((r for r in results if r.expense_id == selected_id), None)
            if selected_result:
                _render_detail(selected_result)

    st.divider()

    # --- Export ---
    exp_col1, exp_col2 = st.columns(2)
    with exp_col1:
        export_filter = st.radio(
            "Include:", ["All results", "Flagged only", "Uncertain only"], horizontal=True
        )
    with exp_col2:
        export_format = st.radio(
            "Format:", ["CSV", "Excel"], horizontal=True
        )

    if export_filter == "Flagged only":
        export_data = flagged
    elif export_filter == "Uncertain only":
        export_data = uncertain
    else:
        export_data = results

    if export_format == "Excel":
        st.download_button(
            "Export Excel",
            data=_export_excel(export_data),
            file_name="audit_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.download_button(
            "Export CSV",
            data=_export_csv(export_data),
            file_name="audit_report.csv",
            mime="text/csv",
        )
    st.caption("This report is a first-pass audit assistant and is not a replacement for human review.")

    st.divider()

    if st.button("New Audit"):
        _reset()


def _reset():
    for key in ["csv_df", "receipt_files", "results", "selected_row",
                "filter_status", "search_query", "processing_step", "selected_row_radio"]:
        st.session_state.pop(key, None)
    st.session_state["step"] = "upload"
    st.rerun()
